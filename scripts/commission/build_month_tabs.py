"""Add the new month's tab to the running SMI Commissions workbook and to the
IPBC Group workbook, following the established layout exactly.

The month is supplied by the caller - there is deliberately no default, because
silently inheriting last month's tab is the exact failure this must never have.

    CM_YEAR=2026 CM_MONTH=5 \
    CM_DIR="c:/Users/ali_m/Downloads/May 2026 Commission" \
    python build_month_tabs.py

The prior month's two workbooks are read from the archive (SRC_DIR) and copied
forward; the new tab is prepended to each and saved under the new month's name.
"""
import json, os, shutil, sys, calendar
import openpyxl

SRC_DIR = os.environ.get(
    "CM_ARCHIVE",
    "C:/Users/ali_m/OneDrive/Business/Colaberry Novedea/Stats/Commissions/SMI Comm")

try:
    YY = int(os.environ["CM_YEAR"])
    MM = int(os.environ["CM_MONTH"])
except (KeyError, ValueError):
    raise SystemExit("FATAL: set CM_YEAR (4-digit) and CM_MONTH (1-12)")
if not 1 <= MM <= 12:
    raise SystemExit("FATAL: CM_MONTH must be 1-12, got %r" % MM)

MON = calendar.month_abbr[MM]                    # "May"
PMM, PYY = (12, YY - 1) if MM == 1 else (MM - 1, YY)
PMON = calendar.month_abbr[PMM]                  # "Apr"

OUT_DIR = os.environ.get("CM_DIR", "c:/Users/ali_m/Downloads/%s %d Commission" % (MON, YY))
DATA_DIR = os.environ.get("CM_DATA_DIR", ".")
TAB = "%s %d" % (MON, YY)

def data(name):
    return os.path.join(DATA_DIR, name)

detail = json.load(open(data("smi_detail.json")))
summary = json.load(open(data("smi_summary.json")))
ipbc = json.load(open(data("ipbc_%s.json" % MON.lower())))

# The pipeline is capped at the reporting month, so its newest rows must BE the
# reporting month. If they are not, the wrong extract is on disk.
newest = max((r["OrderYear"], r["OrderMonth"]) for r in detail)
if newest != (YY, MM):
    raise SystemExit("FATAL: smi_detail.json newest month is %d-%02d, expected %d-%02d"
                     % (newest[0], newest[1], YY, MM))

HDR = ["SMIC_SalesRep", "SMIC_Name", "OrderMonth", "OrderYear", "TotalPaid",
       "AliComm", "SalesRepComm", "MonthlyPaid", "MonthlyPaidMinus25per",
       None, "Name", "OrderMonth", "OrderYear", "Comm"]

# ---------- 1. SMI Commissions ----------
src = "%s/%d_%02d_SMI Commisions.xlsx" % (SRC_DIR, PYY, PMM)
dst = "%s/%d_%02d_SMI Commisions.xlsx" % (OUT_DIR, YY, MM)
if not os.path.exists(src):
    raise SystemExit("FATAL: prior month workbook not in the archive: %s" % src)
shutil.copyfile(src, dst)

wb = openpyxl.load_workbook(dst)
before = list(wb.sheetnames)
if TAB in wb.sheetnames:
    del wb[TAB]
ws = wb.create_sheet(TAB, 0)

for c, h in enumerate(HDR, start=1):
    if h is not None:
        ws.cell(1, c, h)

for i, r in enumerate(detail, start=2):
    ws.cell(i, 1, r["SMIC_SalesRep"])
    ws.cell(i, 2, r["SMIC_Name"])
    ws.cell(i, 3, r["OrderMonth"])
    ws.cell(i, 4, r["OrderYear"])
    ws.cell(i, 5, float(r["TotalPaid"]) if r["TotalPaid"] is not None else None)
    ws.cell(i, 6, float(r["AliComm"]) if r["AliComm"] is not None else None)
    # the export renders a NULL sales-rep commission as the literal string NULL
    ws.cell(i, 7, "NULL" if r["SalesRepComm"] is None else float(r["SalesRepComm"]))
    ws.cell(i, 8, float(r["MonthlyPaid"]) if r["MonthlyPaid"] is not None else None)
    ws.cell(i, 9, float(r["MonthlyPaidMinus25per"]) if r["MonthlyPaidMinus25per"] is not None else None)

for i, r in enumerate(summary, start=2):
    ws.cell(i, 11, r["Name"])
    ws.cell(i, 12, r["OrderMonth"])
    ws.cell(i, 13, r["OrderYear"])
    ws.cell(i, 14, float(r["Comm"]))

wb.save(dst)
print("SMI workbook -> %s" % dst)
print("  copied forward from: %s" % os.path.basename(src))
print("  sheets before: %d   after: %d   new tab first: %s" % (len(before), len(wb.sheetnames), wb.sheetnames[0]))

# ---------- 2. IPBC Group ----------
isrc = "%s/IPBC Group %s %d.xlsx" % (SRC_DIR, PMON, PYY)
idst = "%s/IPBC Group %s %d.xlsx" % (OUT_DIR, MON, YY)
if not os.path.exists(isrc):
    raise SystemExit("FATAL: prior month IPBC workbook not in the archive: %s" % isrc)
shutil.copyfile(isrc, idst)

iwb = openpyxl.load_workbook(idst)
if TAB in iwb.sheetnames:
    del iwb[TAB]
iws = iwb.create_sheet(TAB, 0)

IHDR = ["Customer #", "Source", "Order Date", "Status", "Total", "Amount Paid",
        "Balance Due", "ModifiedDate", "SMIC_SalesRep", "SMIC_CommID", "SMIC_UserID"]
if not ipbc:
    # matches the existing convention for empty months (see the May/Mar/Feb 2024 tabs)
    iws.cell(1, 1, "No Data for the month")
else:
    for c, h in enumerate(IHDR, start=1):
        iws.cell(1, c, h)
    for i, r in enumerate(ipbc, start=2):
        for c, h in enumerate(IHDR, start=1):
            v = r[h]
            iws.cell(i, c, str(v) if h in ("Order Date", "ModifiedDate") and v else v)

iwb.save(idst)
print("IPBC workbook -> %s" % idst)
print("  copied forward from: %s" % os.path.basename(isrc))
print("  sheets: %d   new tab first: %s   rows: %d" % (len(iwb.sheetnames), iwb.sheetnames[0], len(ipbc)))
