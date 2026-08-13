"""Preflight gate for the monthly commission send.

Nothing goes to accounting until this exits 0. It derives both figures from the
files themselves rather than trusting anything typed by hand, refuses to re-send
a month already in the ledger, regression-tests every prior month, and makes a
commission-rate tier change impossible to miss.

    python preflight.py --month 2026-05 --dir "<month folder>"
    python preflight.py --month 2026-05 --dir "<...>" --record   # after a confirmed send

Exit 0 = safe to send, and send_manifest.json is written next to the files.
Exit 1 = do not send. Read the FAIL lines.
"""
import argparse, hashlib, json, os, re, sys, datetime
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
LEDGER = os.path.join(HERE, "ledger.json")

MONTH_ABBR = {1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr", 5: "May", 6: "Jun",
              7: "Jul", 8: "Aug", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec"}

MAX_TOTAL_BYTES = 20 * 1024 * 1024      # comfortably under Mandrill's 25MB ceiling
problems, warnings = [], []


def fail(msg):
    problems.append(msg)
    print("  FAIL  " + msg)


def warn(msg):
    warnings.append(msg)
    print("  WARN  " + msg)


def ok(msg):
    print("  ok    " + msg)


def money(v):
    return "${:,.2f}".format(v)


def expected_names(year, month):
    mon = MONTH_ABBR[month]
    return {
        "original": "%d_%02d_ColaberryTrainingCommissions_Original.xlsx" % (year, month),
        "smi": "%d_%02d_SMI Commisions.xlsx" % (year, month),
        "ipbc": "IPBC Group %s %d.xlsx" % (mon, year),
        "png": "%s %d Staff Commission.png" % (mon, year),
    }


def staff_total(path):
    """Sum column T of the Staff Commissions sheet, the way the emailed figure is derived."""
    wb = openpyxl.load_workbook(path, data_only=True)
    if "Staff Commissions" not in wb.sheetnames:
        fail("'%s' has no 'Staff Commissions' sheet (found: %s)" % (os.path.basename(path), wb.sheetnames))
        return None, []
    ws = wb["Staff Commissions"]
    hdr_a = str(ws.cell(1, 1).value or "").strip().lower()
    hdr_t = str(ws.cell(1, 20).value or "").strip().lower()
    if hdr_a != "name" or hdr_t != "total comm":
        fail("Staff Commissions layout moved: A1=%r T1=%r (expected 'Name' / 'Total Comm'). "
             "Re-read the workbook before trusting any total." % (ws.cell(1, 1).value, ws.cell(1, 20).value))
        return None, []
    rows = []
    for r in range(2, ws.max_row + 1):
        nm = ws.cell(r, 1).value
        if nm is None or str(nm).strip() == "":
            continue
        try:
            rows.append((" ".join(str(nm).split()), float(ws.cell(r, 20).value or 0)))
        except (TypeError, ValueError):
            fail("non-numeric Total Comm for %r at row %d" % (nm, r))
            return None, []
    return round(sum(v for _, v in rows), 2), rows


def smi_figures(path, year, month, tiers):
    """Read the new tab plus the K:N summary block out of the SMI workbook."""
    wb = openpyxl.load_workbook(path, data_only=True)
    tab = "%s %d" % (MONTH_ABBR[month], year)
    if wb.sheetnames[0] != tab:
        fail("SMI workbook's first sheet is %r, expected the new month %r" % (wb.sheetnames[0], tab))
        return None
    ws = wb[tab]

    tp = ac = 0.0
    n = 0
    monthly = set()
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 3).value == month and ws.cell(r, 4).value == year:
            n += 1
            tp += float(ws.cell(r, 5).value or 0)
            ac += float(ws.cell(r, 6).value or 0)
            mp = ws.cell(r, 8).value
            if mp is not None:
                monthly.add(round(float(mp), 2))

    if n == 0:
        fail("SMI tab %r has no rows for %d-%02d" % (tab, year, month))
        return None
    if len(monthly) != 1:
        fail("MonthlyPaid is not uniform across the month's rows: %s" % sorted(monthly))

    summary = {}
    for r in range(2, ws.max_row + 1):
        if str(ws.cell(r, 11).value or "").strip().upper() == "ALI":
            yy, mm, cm = ws.cell(r, 13).value, ws.cell(r, 12).value, ws.cell(r, 14).value
            if yy and mm and cm is not None:
                summary[(int(yy), int(mm))] = round(float(cm), 2)

    if (year, month) not in summary:
        fail("no ALI summary row for %d-%02d in the K:N block" % (year, month))
        return None

    ali = summary[(year, month)]
    if abs(ali - round(ac, 2)) > 0.005:
        fail("K:N summary says %s but the month's AliComm column sums to %s"
             % (money(ali), money(round(ac, 2))))

    company_paid = monthly.pop() if len(monthly) == 1 else round(tp, 2)

    rate = next(t for t in tiers if t["upTo"] is None or company_paid < t["upTo"])
    implied = round(tp * rate["effective"], 2)
    if abs(implied - ali) > 0.02:
        fail("AliComm %s does not match the tier: CompanyPaid %s -> rate %s -> expected ~%s. "
             "Either the tier boundaries moved or the wrong rows were pasted."
             % (money(ali), money(company_paid), rate["rate"], money(implied)))

    return {"tab": tab, "rows": n, "total_paid": round(tp, 2), "ali_comm": ali,
            "company_paid": company_paid, "tier": rate, "summary": summary, "workbook": wb}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--month", required=True, help="YYYY-MM of the commission month")
    ap.add_argument("--dir", required=True, help="folder holding the four artifacts")
    ap.add_argument("--record", action="store_true", help="append this month to the ledger after a confirmed send")
    ap.add_argument("--force", action="store_true", help="override the already-sent guard")
    a = ap.parse_args()

    year, month = (int(x) for x in a.month.split("-"))
    led = json.load(open(LEDGER, encoding="utf-8"))
    tiers = led["tiers"]
    names = expected_names(year, month)
    paths = {k: os.path.join(a.dir, v) for k, v in names.items()}

    print("\nCOMMISSION PREFLIGHT  %s %d" % (MONTH_ABBR[month], year))
    print("=" * 78)

    # ---- 1. already sent? -------------------------------------------------
    print("\n[1] idempotency")
    prior = next((s for s in led["sends"] if s["year"] == year and s["month"] == month), None)
    if prior and not a.record:
        if a.force:
            warn("%s %d was already sent %s - proceeding only because --force was given"
                 % (MONTH_ABBR[month], year, prior.get("sent_utc")))
        else:
            fail("%s %d was ALREADY SENT on %s (staff %s, ali %s). Re-running would double-send "
                 "to accounting. Pass --force only if you intend to send a correction."
                 % (MONTH_ABBR[month], year, prior.get("sent_utc"),
                    money(prior["staff_commission"]), money(prior["ali_comm"])))
    elif not prior:
        ok("no prior send recorded for %s %d" % (MONTH_ABBR[month], year))

    # ---- 2. artifacts present --------------------------------------------
    print("\n[2] attachments")
    total_bytes = 0
    for key, p in paths.items():
        if not os.path.exists(p):
            fail("missing %s: %s" % (key, os.path.basename(p)))
            continue
        sz = os.path.getsize(p)
        total_bytes += sz
        if sz == 0:
            fail("%s is zero bytes" % os.path.basename(p))
        else:
            ok("%-9s %-58s %9d bytes" % (key, os.path.basename(p), sz))
    if total_bytes > MAX_TOTAL_BYTES:
        fail("attachments total %.1f MB, over the %d MB ceiling"
             % (total_bytes / 1048576.0, MAX_TOTAL_BYTES // 1048576))
    else:
        ok("total %.1f MB" % (total_bytes / 1048576.0))

    if problems:
        return finish(a, None, None, None, led, year, month)

    # ---- 3. staff commission ---------------------------------------------
    print("\n[3] staff commission (derived, not typed)")
    staff, rows = staff_total(paths["original"])
    if staff is not None:
        ok("%d mentors, total %s" % (len(rows), money(staff)))
        if any(v < 0 for _, v in rows):
            fail("a mentor has a negative Total Comm")
        big = [(n, v) for n, v in rows if v > 5000]
        for n, v in big:
            warn("unusually large line: %s %s - confirm against Jackie's sheet" % (n, money(v)))

    # ---- 4. ali commission + tier ----------------------------------------
    print("\n[4] ali commission + rate tier")
    smi = smi_figures(paths["smi"], year, month, tiers)
    if smi:
        ok("%s tab: %d rows, TotalPaid %s" % (smi["tab"], smi["rows"], money(smi["total_paid"])))
        ok("CompanyPaid %s -> tier rate %s (effective %.2f%%) -> AliComm %s"
           % (money(smi["company_paid"]), smi["tier"]["rate"],
              smi["tier"]["effective"] * 100, money(smi["ali_comm"])))

        past = [s for s in led["sends"] if s.get("company_paid")]
        if past:
            last = past[-1]
            lt = next(t for t in tiers if t["upTo"] is None or last["company_paid"] < t["upTo"])
            if lt["rate"] != smi["tier"]["rate"]:
                warn("*** RATE TIER CHANGED *** last month (%s %d) sat at %s, this month is %s. "
                     "This is correct per the query, but it is the single easiest thing to get "
                     "wrong by hand - the figure is NOT last month's rate."
                     % (MONTH_ABBR[last["month"]], last["year"], lt["rate"], smi["tier"]["rate"]))
            else:
                ok("same tier as last month (%s)" % lt["rate"])
        head = next((t["upTo"] - smi["company_paid"] for t in tiers
                     if t["upTo"] and smi["company_paid"] < t["upTo"]), None)
        if head is not None and head < 2000:
            warn("only %s below the next tier boundary - a late payment landing in this month "
                 "would change the rate" % money(head))

    # ---- 5. regression against every prior send --------------------------
    print("\n[5] regression - prior months must still compute to what was emailed")
    if smi:
        for s in led["sends"]:
            if (s["year"], s["month"]) == (year, month):
                continue
            got = smi["summary"].get((s["year"], s["month"]))
            if got is None:
                warn("%s %d absent from this workbook's summary block" % (MONTH_ABBR[s["month"]], s["year"]))
            elif abs(got - s["ali_comm"]) > 0.005:
                fail("%s %d was emailed as %s but now computes to %s - the basis was RESTATED. "
                     "Do not send until this is explained."
                     % (MONTH_ABBR[s["month"]], s["year"], money(s["ali_comm"]), money(got)))
            else:
                ok("%s %d still %s" % (MONTH_ABBR[s["month"]], s["year"], money(got)))

    # ---- 6. carried-over tabs intact -------------------------------------
    print("\n[6] carried-over tabs survived the rebuild")
    if smi:
        prev = sorted([s for s in led["sends"] if (s["year"], s["month"]) < (year, month)],
                      key=lambda s: (s["year"], s["month"]))
        prev_tab = "%s %d" % (MONTH_ABBR[prev[-1]["month"]], prev[-1]["year"]) if prev else None
        wb = smi["workbook"]
        if prev_tab and prev_tab in wb.sheetnames:
            ok("%d sheets, previous month's tab %r present" % (len(wb.sheetnames), prev_tab))
        else:
            fail("previous month's tab %r missing from the rebuilt workbook" % prev_tab)
        # Only month tabs matter here. The workbook has always carried a stray
        # empty "Sheet1"; flagging that every month is noise, and noise is how a
        # real signal gets ignored.
        month_tab = re.compile(r"^(%s)\s*\d{4}$" % "|".join(MONTH_ABBR.values()), re.I)
        empties = [s for s in wb.sheetnames
                   if month_tab.match(s.strip()) and wb[s].max_row <= 1 and s != smi["tab"]]
        if empties:
            fail("month tabs came back EMPTY after the rebuild - workbook corruption: %s" % empties[:5])
        else:
            ok("no month tab was emptied by the rebuild")

    return finish(a, staff, smi, paths, led, year, month)


def finish(a, staff, smi, paths, led, year, month):
    print("\n" + "=" * 78)
    if problems:
        print("PREFLIGHT FAILED - %d problem(s), %d warning(s). DO NOT SEND." % (len(problems), len(warnings)))
        for p in problems:
            print("   - " + p)
        return 1

    print("PREFLIGHT PASSED" + ("  (%d warning(s) - read them)" % len(warnings) if warnings else ""))
    for w in warnings:
        print("   ! " + w)

    manifest = {
        "year": year, "month": month,
        "subject": "%s %d Commission (Mentor/Instructor/SMI)" % (MONTH_ABBR[month], year),
        "staff_commission": staff,
        "ali_comm": smi["ali_comm"],
        "company_paid": smi["company_paid"],
        "tier_rate": smi["tier"]["rate"],
        "to": ["accounting@colaberry.com"],
        "bcc": ["ali@colaberry.com"],
        "attachments": [],
        "generated_utc": datetime.datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
    }
    for key in ("original", "smi", "ipbc", "png"):
        p = paths[key]
        h = hashlib.sha256(open(p, "rb").read()).hexdigest()
        manifest["attachments"].append({"role": key, "filename": os.path.basename(p),
                                        "bytes": os.path.getsize(p), "sha256": h})

    mpath = os.path.join(a.dir, "send_manifest.json")
    json.dump(manifest, open(mpath, "w", encoding="utf-8"), indent=2)
    print("\n  Staff Commission: %s" % money(manifest["staff_commission"]))
    print("  Ali Commission:   %s" % money(manifest["ali_comm"]))
    print("\n  manifest -> %s" % mpath)

    if a.record:
        if any(s["year"] == year and s["month"] == month for s in led["sends"]):
            print("  ledger already contains this month - nothing to record")
        else:
            led["sends"].append({
                "year": year, "month": month, "subject": manifest["subject"],
                "staff_commission": manifest["staff_commission"],
                "ali_comm": manifest["ali_comm"],
                "company_paid": manifest["company_paid"],
                "sent_utc": datetime.datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
            })
            led["sends"].sort(key=lambda s: (s["year"], s["month"]))
            json.dump(led, open(LEDGER, "w", encoding="utf-8"), indent=2)
            print("  ledger updated - %s %d recorded" % (MONTH_ABBR[month], year))
    else:
        print("\n  Next: send, then re-run with --record to close the loop.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
