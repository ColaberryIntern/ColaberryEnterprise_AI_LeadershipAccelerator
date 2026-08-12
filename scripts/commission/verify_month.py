"""Independent read-back of the built April workbooks — verify against the
known-good March file rather than against the JSON they were written from."""
import openpyxl

OUT = "c:/Users/ali_m/Downloads/April 2026 Commission"
CANON = "C:/Users/ali_m/OneDrive/Business/Colaberry Novedea/Stats/Commissions/SMI Comm"

wb = openpyxl.load_workbook(OUT + "/2026_04_SMI Commisions.xlsx", data_only=True)
ws = wb["Apr 2026"]
print("=== 2026_04_SMI Commisions.xlsx ===")
print("sheets: %d   first: %s   second: %s" % (len(wb.sheetnames), wb.sheetnames[0], wb.sheetnames[1]))
print("header:", [ws.cell(1, c).value for c in range(1, 15)])

apr_tp = apr_ac = 0.0
apr_n = 0
monthly = set()
for r in range(2, ws.max_row + 1):
    if ws.cell(r, 3).value == 4 and ws.cell(r, 4).value == 2026:
        apr_n += 1
        apr_tp += float(ws.cell(r, 5).value or 0)
        apr_ac += float(ws.cell(r, 6).value or 0)
        monthly.add(round(float(ws.cell(r, 8).value or 0), 2))
print("Apr 2026: rows=%d  TotalPaid=%.2f  AliComm=%.4f  MonthlyPaid=%s" % (apr_n, apr_tp, apr_ac, monthly))
print("TotalPaid x 0.1125 = %.4f  -> rounded $%s" % (apr_tp * 0.1125, format(round(apr_ac, 2), ',.2f')))

summ = {}
for r in range(2, ws.max_row + 1):
    if ws.cell(r, 11).value == "ALI":
        summ[(ws.cell(r, 13).value, ws.cell(r, 12).value)] = float(ws.cell(r, 14).value)
print("\nsummary block (K:N) recent:")
for k in sorted(summ, reverse=True)[:6]:
    print("   %d-%02d  %.2f" % (k[0], k[1], summ[k]))

# regression: prior months in the NEW file must equal what was emailed
EMAILED = {(2026, 3): 2143.44, (2026, 2): 1907.19, (2025, 12): 3768.11, (2025, 11): 1177.93}
print("\nregression vs what was actually emailed:")
ok = True
for k, v in sorted(EMAILED.items(), reverse=True):
    got = summ.get(k)
    good = got is not None and abs(got - v) < 0.005
    ok = ok and good
    print("   %d-%02d  emailed=%.2f  file=%s  %s" % (k[0], k[1], v, got, "OK" if good else "MISMATCH"))

# the untouched March tab must survive the round-trip intact
old = openpyxl.load_workbook(CANON + "/2026_03_SMI Commisions.xlsx", data_only=True)
for tab in ("Mar 2026", "Feb 2026", "Jan 2026"):
    a, b = wb[tab], old[tab]
    same = a.max_row == b.max_row and a.max_column == b.max_column
    tot_a = sum(float(a.cell(r, 5).value or 0) for r in range(2, a.max_row + 1))
    tot_b = sum(float(b.cell(r, 5).value or 0) for r in range(2, b.max_row + 1))
    print("   carried-over tab %-9s rows %s/%s  TotalPaid %.2f vs %.2f  %s"
          % (tab, a.max_row, b.max_row, tot_a, tot_b,
             "OK" if same and abs(tot_a - tot_b) < 0.005 else "CHANGED"))

iwb = openpyxl.load_workbook(OUT + "/IPBC Group Apr 2026.xlsx", data_only=True)
iws = iwb["Apr 2026"]
print("\n=== IPBC Group Apr 2026.xlsx ===")
print("sheets: %d   first: %s   A1: %r" % (len(iwb.sheetnames), iwb.sheetnames[0], iws.cell(1, 1).value))
print("Mar 2026 tab still present with %d rows" % iwb["Mar 2026"].max_row)
print("\nALL PRIOR-MONTH REGRESSIONS PASS" if ok else "\n*** REGRESSION FAILURE ***")
